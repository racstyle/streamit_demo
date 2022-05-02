import streamlit as st
import pandas as pd
import openpyxl
import datetime

# Helper function to init session state to use in pages
def init_session_state(session_name, init_val):
    if session_name not in st.session_state:    # if a 
        st.session_state[session_name] = init_val
    return


def app():
    st.title("Example: Order form")
    st.write("Now that we know the basics of Streamlit, let's see an example of all these components playing together.")
    st.write("The following app is similar to the web app I've done in my client project.")
    st.markdown("---")
    
    # Using the same sample Excel data form the previous page
    excel_file = pd.read_excel("example_excel_data.xlsx", sheet_name="Sheet1", engine="openpyxl")  # read in Excel file via pandas library
    excel_dataframe = pd.DataFrame(excel_file)    # converting raw data into pandas dataframe to make it easy to read and manipulate in Python
    excel_dataframe["Order Date"] = excel_dataframe["Order Date"].dt.strftime("%m/%d/%Y")   # convert date in Excel to Python-friendly DataFrame
    
    # Creating a simple form to "order" supplies and see it in the Excel sheet
    st.header("Order form")
    
    # Choose the number of people in this form
    max_orders = 5      # max number of orders allowed
    num_orders = st.selectbox(label="Select the number of orders, max is 5", options=list(range(1, max_orders+1)), key="num_orders")     # options will generate a list of numbers counting up fr 1-max_orders, key parameter is used for adding statefulness
    
    # Number of rows
    form_rows = [0]*max_orders
    
    # Today's date + list for summary table
    today_date = datetime.date.today()     # today's date
    date_list = [today_date.strftime("%m/%d/%Y")]*max_orders
    
    # Empty lists of inputs, done so that we can fill the form in any order
    form_region = [""]*max_orders
    form_name = [""]*max_orders
    form_item = [""]*max_orders
    form_units = [0]*max_orders
    
    # List of unique values in regions and items
    region_list = excel_dataframe["Region"].unique()   # this allows me to get a list of unique values in the given Excel/dataframe column
    items_list = excel_dataframe["Item"].unique()   # this allows me to get a list of unique values in the given Excel/dataframe column
    
    # Mapping pricing of each item
    item_pricing = {
        "Binder": 3.99,
        "Pencil": 1.40,
        "Pen": 1.50,
        "Paper": 1.29,
        "Pen Set": 5.29
    }
    
    # Empty lists of each item price + total price for each name
    item_price = [0]*max_orders
    total_price = [0]*max_orders
    
    # Session state names for showing inputs in summary page
    state_region = "region_state"
    state_name = "name_state"
    state_item = "item_state"
    state_units = "units_state"
    
    
    # We append number of "layers"/orders according to num_orders
    for i in range(0, num_orders + 1):
        # First we create some columns for the region, name (or name), item, and # units
        form_cols = st.columns((2, 4, 4, 3, 3))
        
        # Header row
        if i == 0:
            form_cols[0].write("Row #")
            form_cols[1].write("Region")
            form_cols[2].write("Name")
            form_cols[3].write("Item")
            form_cols[4].write("Number of units")
        
        # Actual form, this allows to create a new row each iteration
        if i > 0:
            # Row number
            form_rows[i-1] = i
            form_cols[0].write("")      # extra space to line up with inputs
            form_cols[0].write("")      # extra space to line up with inputs
            form_cols[0].write(str(form_rows[i-1]))
            
            # Inputs
            form_region[i-1] = form_cols[1].selectbox(label="", options=region_list, key=state_region + "_layer_" + str(i) + "_input")  # region
            form_name[i-1] = form_cols[2].text_input(label="", placeholder="Enter your name", key=state_name + "_layer_" + str(i) + "_input")   # name, label param is a requirement even if you have nothing to put in there
            form_item[i-1] = form_cols[3].selectbox(label="", options=items_list, key=state_item + "_layer_" + str(i) + "_input")      # item
            form_units[i-1] = form_cols[4].number_input(label="", step=1, format="%d", key=state_units + "_layer_" + str(i) + "_input")     # units/how many, step allows you to increment by 1 and format removes leading zeros
            
            # Calculate totals for each name
            item_price[i-1] = item_pricing[form_item[i-1]]          # get unit pricing
            total_price[i-1] = form_units[i-1] * item_price[i-1]    # get total pricing for each name
    
    
    # Submit button
    if st.button("Submit"):
        st.markdown("---")
        
        error_flag = False      # flag for error checking
        
        # Going through each row of form
        for i in range(0, num_orders):      # for each row
            # Error checking
            if form_name[i] == "" or form_units[i] <= 0:
                error_msg = "Error in **row " + str(i+1) + "**!"    # shows which row has invalid input
                
                # Name is blank
                if form_name[i] == "": error_msg += "  The **name field** is blank!"
                # Number of units is 0 or below
                else: error_msg += "  The **number of units** is zero or negative!"
                
                # Display error message
                st.error(error_msg)
                
                # Flag that form has error so it won't submit the form and exit for loop
                error_flag = True
                break
            
            # Init input session states to see inputs in summary page
            init_session_state(state_region + "_layer_" + str(i), form_region[i-1])   # region session state
            init_session_state(state_name + "_layer_" + str(i), form_name[i-1])         # name session state
            init_session_state(state_item + "_layer_" + str(i), form_item[i-1])       # item session state
            init_session_state(state_units + "_layer_" + str(i), form_units[i-1])     # units session state
            
            # Update input session states when form changes
            st.session_state[state_region + "_layer_" + str(i)] = form_region[i-1]
            st.session_state[state_name + "_layer_" + str(i)] = form_name[i-1]
            st.session_state[state_item + "_layer_" + str(i)] = form_item[i-1]
            st.session_state[state_units + "_layer_" + str(i)] = form_units[i-1]
            
            # Need to open Excel file with different library to be able to write in Excel
            excel_file_write = openpyxl.load_workbook("example_excel_data.xlsx")    # open Excel file with different Python library
            excel_file_write_active = excel_file_write.active   # enable opened Excel to be modified by Python
            
            # Writing into the Excel file
            excel_file_write_active.insert_rows(idx=2)      # insert an empty row for each num_orders in Excel file below the header row
            excel_file_write_active.cell(row=2, column=1, value=today_date.strftime("%m/%d/%Y"))         # today's date
            excel_file_write_active.cell(row=2, column=2, value=form_region[i])     # region
            excel_file_write_active.cell(row=2, column=3, value=form_name[i])       # name
            excel_file_write_active.cell(row=2, column=4, value=form_item[i])       # item
            excel_file_write_active.cell(row=2, column=5, value=form_units[i])      # units
            excel_file_write_active.cell(row=2, column=6, value=item_pricing[form_item[i]])                     # individual item price
            excel_file_write_active.cell(row=2, column=7, value=form_units[i] * item_pricing[form_item[i]])     # total price = indiv item price * num units
            excel_file_write.save("example_excel_data.xlsx")    # save the Excel file for each row
        
        # Do not continue unless the form is valid
        if error_flag == False:
            # Set up dataframe table to show current input
            submitted_header = ["Order Date", "Region", "Name", "Item", "Units", "Item Price", "Total Cost"]     # header for output table
            submitted_form = pd.DataFrame(
                list(zip(
                    date_list[:num_orders],     # today's date
                    form_region[:num_orders],   # regions
                    form_name[:num_orders],     # names
                    form_item[:num_orders],     # items
                    form_units[:num_orders],    # units
                    item_price[:num_orders],    # price of selected item
                    total_price[:num_orders],   # total price of selected item with # units
                )),
                columns=submitted_header
            )
            submitted_form = submitted_form
            
            # Totals row
            totals = pd.DataFrame(
                [[
                "TOTALS", "", "", "",       # order date with TOTALS, skipping region, name, and item columns
                    round(sum(form_units)),     # format units to not show decimals
                    0,                          # due to number formatting, leave this 0
                    sum(total_price)            # format total price to show 2 decimal places
                ]],
                columns=submitted_header
            )
            submitted_form_summary = submitted_form.append(totals, ignore_index=True)
            
            # Showing submitted form
            submitted_cols = st.columns((7, 3))             # 2 columns
            submitted_cols[0].table(submitted_form_summary.style.format({submitted_header[4]: "{:.0f}", submitted_header[5]: "{:.2f}", submitted_header[6]: "{:.2f}"}))     # showing output table with formatting
            submitted_cols[1].write("Total items: " + str(round(sum(form_units))))    # total number of items
            submitted_cols[1].write("Total price: $" + str(sum(total_price)))    # total price of items * units
            
            
            # Session states for summary page
            # Recent order summary
            init_session_state("recent_order_table", [])    # init session state for recent order in next page
            st.session_state["recent_order_table"] = submitted_form.style.format({submitted_header[4]: "{:.0f}", submitted_header[5]: "{:.2f}", submitted_header[6]: "{:.2f}"})
            
            # Add recent order to all orders for next page
            all_orders_dataframe = excel_dataframe      # make copy of Excel dataframe to preserve original
            all_orders_dataframe = submitted_form.append(all_orders_dataframe, ignore_index=True)   # prepend recent order to all orders
            init_session_state("all_orders_table", [])         # init session state for summary in next page
            st.session_state["all_orders_table"] = all_orders_dataframe.style.format({submitted_header[4]: "{:.0f}", submitted_header[5]: "{:.2f}", submitted_header[6]: "{:.2f}"})     # add table to summary session state
            
            
            # Success message
            st.write("")    # extra space
            st.success("Done!  Thank you for your order!  Go to the next page to see your order summary.")
            
    return